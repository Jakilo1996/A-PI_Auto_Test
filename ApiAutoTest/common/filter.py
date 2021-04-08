# -*- coding:utf8 -*- #
# -----------------------------------------------------------------------------------
# ProjectName:   test_demo
# FileName:     filter.py
# Author:      Jakiro
# Datetime:    2021/3/18 4:42 下午
# Description:
# 命名规则  文件名小写字母+下划线，类名大驼峰，方法、变量名小写字母+下划线连接
# 常量大写，变量和常量用名词、方法用动词
# -----------------------------------------------------------------------------------

# -*- coding:utf8 -*- #
# -----------------------------------------------------------------------------------
# ProjectName:   test_demo
# FileName:     priority_filter.py
# Author:      Jakiro
# Datetime:    2021/3/18 3:11 下午
# Description:   用于将excel文件中的sheet中的优先级文件
# 命名规则  文件名小写字母+下划线，类名大驼峰，方法、变量名小写字母+下划线连接
# 常量大写，变量和常量用名词、方法用动词
# -----------------------------------------------------------------------------------


import openpyxl


# 去取excel文件模块 ，输出格式为列表套元组
def read_excel_data(file_path, all_sheet=True, sheet=None):
    # 打开表格  实例化workbook对象 将文件路径作为参数
    get_wb = openpyxl.load_workbook(file_path)
    if all_sheet:
        # 获取全部的sheet页名
        sheet_lists = get_wb.sheetnames
        all_lists = []
        for sheet in sheet_lists:
            # 指定sheet 页 默认为第一个
            get_sheet = get_wb[sheet]
            one_line_data_list = []
            stats = True
            # 循环遍历每一行的tuple对象
            for row_tuple in get_sheet:
                if stats:
                    stats = False
                    continue
                # 循环遍历每个单元格
                for cell in row_tuple:
                    # 拼接 value
                    one_line_data_list.append(cell.value)
                    # 转化成元组
                    one_line_data_tuple = tuple(one_line_data_list)
                # 清空临时列表
                one_line_data_list.clear()
                # 将临时元组添加到总列表中
                all_lists.append(one_line_data_tuple)
        return all_lists
    else:
        if sheet:
            all_lists = []
            get_sheet = get_wb[sheet]
            one_line_data_list = []
            stats = True
            # 循环遍历每一行的tuple对象
            for row_tuple in get_sheet:
                if stats:
                    stats = False
                    continue
                # 循环遍历每个单元格
                for cell in row_tuple:
                    # 拼接 value
                    one_line_data_list.append(cell.value)
                    # 转化成元组
                    one_line_data_tuple = tuple(one_line_data_list)
                # 清空临时列表
                one_line_data_list.clear()
                # 将临时元组添加到总列表中
                all_lists.append(one_line_data_tuple)
        return all_lists


def priority_filter(file_path, new_file_path, filter_field, col, sheet_name=None, all_sheet=True):
    pre_data_lists = read_excel_data(file_path, all_sheet=all_sheet, sheet=sheet_name)
    data_lists = []
    for pre_data_list in pre_data_lists:
        if pre_data_list[col] == filter_field:
            data_lists.append(pre_data_list)
    new_workbook = openpyxl.Workbook()
    new_worksheet = new_workbook.create_sheet('new_sheet')
    for row in data_lists:
        new_worksheet.append(row)
    new_workbook.save(new_file_path)


if __name__ == '__main__':
    print(priority_filter(r'/Users/qiujie/PycharmProjects/test_demo/excel/api_auto_template.xlsx',
                          r'/Users/qiujie/PycharmProjects/test_demo/excel/copy.xlsx', 'Y', 4))

# -*- coding:utf8 -*- #
# -----------------------------------------------------------------------------------
# ProjectName:   ApiAutoTest
# FileName:     read_excel.py
# Author:      Jakiro
# Datetime:    2021/3/5 11:31 上午
# Description: 此方法通过openpyxl模块，实现用例相关数据文件的读取
# 命名规则  文件名小写字母+下划线，类名大驼峰，方法、变量名小写字母+下划线连接
# 常量大写，变量和常量用名词、方法用动词
# -----------------------------------------------------------------------------------

import openpyxl
from common.read_json import read_json
from common.read_ini import ReadIni
from data.project_constant import ProjectConstant


class ReadExcel(object):
    def __init__(self):
        # 实例化ReadIni类的一个对象
        self.read_ini = ReadIni()
        # 通过调用read_ini的指定方法获得excel的数据
        self.excel_data = self.read_ini.get_excel_path()
        # 通过调用read_ini的指定方法获得参数的json数据路径
        self.params_data_path = self.read_ini.get_params_json_path()
        # 通过调用read_ini的指定方法获得期望结果的json数据路径
        self.exception_data_path = self.read_ini.get_exception_json_path()
        # 通过openpyxl的load_workbook方法实例化一个workbook对象
        self.wb = openpyxl.load_workbook(self.excel_data)
        # 通过workbook对象本质是一个字典 通过表名key去 获得一个worksheet对象 通过readini的getsheet方法获得ini文件中的键名 也就是key
        self.ws = self.wb[self.read_ini.get_sheet_name()]

    # 声明一个方法获得指定单元格的value
    def get_cell_value(self, row, column):
        # 通过ws对象的列行坐标找到对应的单元格 然后通过cell对象的，value属性获取指定单元格的值
        return self.ws[column + str(row)].value

    # 声明一个方法获得用例编号
    def get_case_id(self, row):
        # 通过ProjectConstant类的属性获得对应的常量值
        # p#rint(type(ProjectConstant.CASE_ID))
        return self.get_cell_value(row, ProjectConstant.CASE_ID)

    # 声明一个方法获得用例标题
    def get_case_name(self, row):
        # 通过ProjectConstant类的属性获得对应的常量值
        return self.get_cell_value(row, ProjectConstant.CASE_NAME)

    # 声明一个方法获得用例的请求方法
    def get_case_method(self, row):
        # 通过ProjectConstant类的属性获得对应的常量值
        return self.get_cell_value(row, ProjectConstant.REQUEST_METHOD)

    # 声明一个方法获得用例的请求的url
    def get_case_url(self, row):
        # 通过ProjectConstant类的属性获得对应的常量值
        return self.get_cell_value(row, ProjectConstant.REQUEST_URL)

    # 声明一个方法获得用例请求的前置条件
    def get_case_precondition_id(self, row):
        # 通过ProjectConstant类的属性获得对应的常量值
        return self.get_cell_value(row, ProjectConstant.PRECONDITION_ID)

    # 声明一个方法获得用例请求的依赖字段
    def get_case_depend_field(self, row):
        # 通过ProjectConstant类的属性获得对应的常量值
        return self.get_cell_value(row, ProjectConstant.DEPEND_FIELD)

    # 声明一个方法获得用例执行需要的正则表达式
    def get_case_regular_expression(self, row):
        # 通过ProjectConstant类的属性获得对应的常量值
        return self.get_cell_value(row, ProjectConstant.REGULAR_EXPRESSION)

    # 声明一个方法或的用例执行需要的参数的key
    def get_case_parameters_key(self, row):
        # 通过ProjectConstant类的属性获得对应的常量值
        return self.get_cell_value(row, ProjectConstant.REQUEST_PARAMETERS)

    # 声明一个方法或的用例执行的预期结果的key
    def get_case_expected_outcome_key(self, row):
        # 通过ProjectConstant类的属性获得对应的常量值
        return self.get_cell_value(row, ProjectConstant.EXPECTED_OUTCOME)

    # 声明一个方法获得用例执行需要参数的值
    def get_case_parameters_value(self, row):
        # 通过read_json方法获得指定json文件，然后根据key获得请求所需参数的值
        if self.get_case_parameters_key(row):
            return read_json(self.params_data_path)[self.get_case_parameters_key(row)]

    # 声明一个方法获得用例执行的期望结果的值
    def get_case_expected_outcome_value(self, row):
        # 通过read_json方法获得指定json文件，然后根据key获得期望结果的value
        if self.get_case_expected_outcome_key(row):
            return read_json(self.exception_data_path)[self.get_case_expected_outcome_key(row)]

    # 声明一个方法获得用例执行的响应数据类型
    def get_case_response_data_type(self, row):
        # print(self.get_cell_value(row,ProjectConstant.RESPONSE_DATA_TYPE))
        return self.get_cell_value(row, ProjectConstant.RESPONSE_DATA_TYPE)

    # 声明一个方法 获得用例是否执行参数
    def get_case_if_execute(self, row):
        return self.get_cell_value(row, ProjectConstant.CASE_IF_EXECUTE)

    # 声明一个方法获取excel的行数
    def get_row_count(self):
        # 通过work_sheet对象的max_row属性 获得行数
        return self.ws.max_row

    # 声明一个方法获取excel的列数
    def get_column_count(self):
        # 通过work_sheet对象的max_column属性 获得列数
        return self.ws.max_column


if __name__ == '__main__':
    read_excel = ReadExcel()
    print(read_excel.get_cell_value(2, 'C'))
    print(read_excel.get_case_id(2))
    print(read_excel.get_case_name(2))
    print(read_excel.get_case_method(2))
    print(read_excel.get_case_url(2))
    print(read_excel.get_case_if_execute(2))
    print(read_excel.get_case_precondition_id(8))
    print(read_excel.get_case_depend_field(7))
    print(read_excel.get_case_regular_expression(7))
    print(read_excel.get_case_parameters_key(4))
    print(read_excel.get_case_expected_outcome_key(4))
    print(read_excel.get_case_parameters_value(4))
    print(read_excel.get_case_expected_outcome_value(4))
    print(read_excel.get_case_response_data_type(4))
    print(read_excel.get_row_count())
    print(read_excel.get_column_count())

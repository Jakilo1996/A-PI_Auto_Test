# -*- coding:utf8 -*- #
# -----------------------------------------------------------------------------------
# ProjectName:   ranzhi_project
# FileName:     excel_to_yaml
# Author:      Jakiro
# Datetime:    2020/12/8 14:36
# Description:
# -----------------------------------------------------------------------------------


# 将excel文件转化为yaml文件

import yaml
from common.read_excel import read_excel_data
from common.raed_yaml import read_yaml_data
import openpyxl

from openpyxl import load_workbook
import yaml

from common.read_excel import read_excel_data

# 转化excel格式为yaml格式 excel中的空单元格 用null填充
def excel_to_yaml(path, yaml_name):
    excel_data = load_workbook(path)
    with open(yaml_name, mode='w', encoding='utf8') as fi:
        # 获取sheet
        # repr 把参数转为字符串类型 给两边直接加''
        for sheet in excel_data.sheetnames:
            fi.write(repr(sheet) + ':\n')
            get_sheet = excel_data[sheet]
            data_list = []
            header_filter = True
            # 遍历每行，取得每行值的元组
            for row_tuple in get_sheet:
                # 过滤表头
                if header_filter:
                    header_filter = False
                    continue
                # 用列表推导式将每行数据以列表形式添加到data_list中（遍历每行元组，取得单元格cell值）
                line = '  - ' + str([str(cell.value) for cell in row_tuple]) + '\n'
                fi.write(line)


if __name__ == '__main__':
    # path = r'C:\Users\Administrator\codes\test\project\ranzhi_project\data_config\login_data.xlsx'
    # yaml_name = r'C:\Users\Administrator\codes\test\project\ranzhi_project\data_config\login_test_data1.yaml'
    # excel_to_yaml(path, yaml_name)
    path = r'C:\Users\Administrator\codes\test\project\ranzhi_project\data_config\useradd_data.xlsx'
    yaml_name = r'C:\Users\Administrator\codes\test\project\ranzhi_project\data_config\useradd_test_data.yaml'
    excel_to_yaml(path, yaml_name)

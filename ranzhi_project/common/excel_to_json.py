# -*- coding:utf8 -*- #
# -----------------------------------------------------------------------------------
# ProjectName:   ranzhi_project
# FileName:     excel_to_json
# Author:      Jakiro
# Datetime:    2020/12/10 14:14
# Description:
# -----------------------------------------------------------------------------------

import os
from openpyxl import load_workbook
from common.read_ini import ReadIni

# 转化excel格式为json格式

def dict_to_json(json_path, json_dict):
    with open(json_path, mode='w', encoding='utf8') as fi:
        fi.write('{\n')
        length1 = len(json_dict)
        i = 0
        keys = list(json_dict.keys())
        while i < length1:
            # 获取key值
            fi.write(' ' * 2 + '"' + str(keys[i]) + '"' + ':\n')
            fi.write(' ' * 2 + '[\n')
            elem = json_dict[keys[i]]
            length2 = len(elem)
            for j in range(length2):
                fi.write(' ' * 4 + '[')
                row = elem[j]
                length3 = len(row)
                for z in range(length3):
                    cell = row[z]
                    if z < length3 - 1:
                        fi.write('"' + str(cell) + '", ')
                    else:
                        fi.write('"' + str(cell) + '"')
                if j < length2 - 1:
                    fi.write('],\n')
                else:
                    fi.write(']\n')
            if i < length1 - 1:
                fi.write(' ' * 2 + '],\n')
            else:
                fi.write(' ' * 2 + ']\n')
            i += 1
        fi.write('}\n')


def excel_to_json(excel_path, json_path, sheets=[]):
    if os.path.isfile(json_path):
        with open(json_path, mode='r', encoding='utf8') as fi:
            try:
                json_dict = eval(fi.read())
            except:
                json_dict = {}
        excel_data = load_workbook(excel_path)
        if type(sheets) == str:
            sheet_seq = [sheets]
        elif type(sheets) == list or type(sheets) == tuple:
            if sheets:
                sheet_seq = sheets
            else:
                sheet_seq = excel_data.sheetnames
        else:
            raise TypeError("输入的sheets不是可迭代对象，请输入字符串或列表或元组")
        for sheet in sheet_seq:
            get_sheet = excel_data[sheet]
            row_list = []
            header_filter = True
            for row_tuple in get_sheet:
                if header_filter:
                    header_filter = False
                    continue
                row_list.append([str(cell.value) for cell in row_tuple])
            # json_dict[sheet+'1'] = row_list
            json_dict[sheet] = row_list
        dict_to_json(json_path, json_dict)
    else:
        with open(json_path, mode='w', encoding='utf8') as fi:
            pass
        excel_to_json(excel_path, json_path)


if __name__ == '__main__':
    json_path = 'testcase_data.json'
    excel_to_json(ReadIni().case_excel_abspath(), json_path)
